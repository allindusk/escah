import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

BASE = r"d:\D11_DeveloperProject\150_HTML_Project\escalation_heroines\escah\tools\_texts_for_translation"
OUT = r"d:\D11_DeveloperProject\150_HTML_Project\escalation_heroines\escah\tools\_char_counts_new.xlsx"

KEY_RE = re.compile(r'^\s*\[\d+\]')

rows = []
for root, _, files in os.walk(BASE):
    for f in files:
        if not f.lower().endswith(".txt"):
            continue
        p = os.path.join(root, f)
        try:
            t = open(p, encoding="utf-8").read()
        except UnicodeDecodeError:
            t = open(p, encoding="utf-8-sig").read()
        total = len(t)
        visible = sum(1 for c in t if not c.isspace())
        keys = sum(1 for line in t.splitlines() if KEY_RE.match(line))
        rel = os.path.relpath(p, BASE)
        group = "characters" if rel.startswith("characters" + os.sep) else "(toplevel)"
        rows.append((rel, group, total, visible, keys))

# 默认按文件夹内顺序：先排完所有 (toplevel)，再排 characters/ 子目录
GROUP_ORDER = {"(toplevel)": 0, "characters": 1}
rows.sort(key=lambda r: (GROUP_ORDER.get(r[1], 9), r[0]))

wb = Workbook()

# ---- Sheet 1: 全文件明细（文件夹顺序）----
ws = wb.active
ws.title = "明细"
hdr = ["文件", "分组", "总字符数(含空白)", "可见字符数(不含空白)", "Key数量([N])"]
ws.append(hdr)
for c in range(1, len(hdr) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="305496")
    cell.alignment = Alignment(horizontal="center")
for rel, group, total, visible, keys in rows:
    ws.append([rel, group, total, visible, keys])
ws.append(["TOTAL", "", sum(r[2] for r in rows), sum(r[3] for r in rows), sum(r[4] for r in rows)])
last = ws.max_row
for c in range(1, len(hdr) + 1):
    ws.cell(row=last, column=c).font = Font(bold=True)
ws.column_dimensions["A"].width = 48
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 14
ws.freeze_panes = "A2"

# ---- Sheet 2: 按分组汇总 ----
ws2 = wb.create_sheet("分组汇总")
ws2.append(["分组", "文件数", "总字符数", "可见字符数", "Key数量([N])"])
for c in range(1, 6):
    cell = ws2.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="548235")
    cell.alignment = Alignment(horizontal="center")
groups = {}
for rel, group, total, visible, keys in rows:
    g = groups.setdefault(group, [0, 0, 0, 0])
    g[0] += 1
    g[1] += total
    g[2] += visible
    g[3] += keys
for gname in sorted(groups):
    n, tot, vis, k = groups[gname]
    ws2.append([gname, n, tot, vis, k])
ws2.append(["TOTAL", len(rows), sum(r[2] for r in rows), sum(r[3] for r in rows), sum(r[4] for r in rows)])
for c in range(1, 6):
    ws2.cell(row=ws2.max_row, column=c).font = Font(bold=True)
ws2.column_dimensions["A"].width = 16
for col in ("B", "C", "D", "E"):
    ws2.column_dimensions[col].width = 18

wb.save(OUT)
print(f"{len(rows)} 文件 → {OUT}")
print(f"全站总字符(含空白)={sum(r[2] for r in rows)}  可见={sum(r[3] for r in rows)}  Key总数={sum(r[4] for r in rows)}")
for gname in sorted(groups):
    n, tot, vis, k = groups[gname]
    print(f"  {gname}: {n} 文件, 总 {tot}, 可见 {vis}, Key {k}")
